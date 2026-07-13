from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from sqlalchemy.orm import Session, joinedload

from app.models.attendance import AttendanceRecord
from app.models.employee import Employee
from app.schemas.report import ReportRow


class ReportService:
    def _build_rows(self, records: list[AttendanceRecord]) -> list[ReportRow]:
        rows: list[ReportRow] = []
        for record in records:
            full_name = " ".join(
                part.strip() for part in [record.employee.first_name, record.employee.last_name] if part and part.strip()
            )
            rows.append(
                ReportRow(
                    employee_code=record.employee.employee_code,
                    employee_name=full_name,
                    department=record.employee.department.name if record.employee.department else None,
                    job_title=record.employee.job_title,
                    attendance_date=record.attendance_date.isoformat(),
                    check_in_time=record.check_in_time.isoformat() if record.check_in_time else None,
                    check_out_time=record.check_out_time.isoformat() if record.check_out_time else None,
                    working_hours=round(record.working_hours, 2),
                    status=record.status,
                    is_late=record.is_late,
                )
            )
        return rows

    def daily_report(self, db: Session, report_date: date, branch_id: int | None = None) -> list[ReportRow]:
        records = (
            db.query(AttendanceRecord)
            .options(joinedload(AttendanceRecord.employee).joinedload(Employee.department))
            .filter(AttendanceRecord.attendance_date == report_date)
        )
        if branch_id:
            records = records.join(Employee).filter(Employee.branch_id == branch_id)
        records = records.order_by(AttendanceRecord.id.desc()).all()
        return self._build_rows(records)

    def weekly_report(self, db: Session, report_date: date, branch_id: int | None = None) -> list[ReportRow]:
        week_start = report_date - timedelta(days=report_date.weekday())
        week_end = week_start + timedelta(days=7)
        
        records = (
            db.query(AttendanceRecord)
            .options(joinedload(AttendanceRecord.employee).joinedload(Employee.department))
            .filter(AttendanceRecord.attendance_date >= week_start, AttendanceRecord.attendance_date < week_end)
        )
        if branch_id:
            records = records.join(Employee).filter(Employee.branch_id == branch_id)
        records = records.order_by(AttendanceRecord.attendance_date.desc(), AttendanceRecord.id.desc()).all()
        return self._build_rows(records)

    def monthly_report(self, db: Session, month: str, branch_id: int | None = None) -> list[ReportRow]:
        month_start = datetime.strptime(f"{month}-01", "%Y-%m-%d").date()
        if month_start.month == 12:
            month_end = date(month_start.year + 1, 1, 1)
        else:
            month_end = date(month_start.year, month_start.month + 1, 1)

        records = (
            db.query(AttendanceRecord)
            .options(joinedload(AttendanceRecord.employee).joinedload(Employee.department))
            .filter(AttendanceRecord.attendance_date >= month_start, AttendanceRecord.attendance_date < month_end)
        )
        if branch_id:
            records = records.join(Employee).filter(Employee.branch_id == branch_id)
        records = records.order_by(AttendanceRecord.attendance_date.desc(), AttendanceRecord.id.desc()).all()
        return self._build_rows(records)

    def export_to_excel(self, report_data: list[ReportRow]) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الحضور"
        
        # Header
        headers = ["كود الموظف", "اسم الموظف", "القسم", "المسمى الوظيفي", "تاريخ الحضور", "وقت الحضور", "وقت الانصراف", "ساعات العمل", "الحالة", "متأخر"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for row in report_data:
            ws.append([
                row.employee_code,
                row.employee_name,
                row.department,
                row.job_title,
                row.attendance_date,
                row.check_in_time,
                row.check_out_time,
                row.working_hours,
                row.status,
                "نعم" if row.is_late else "لا"
            ])
        
        # Auto-fit columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_to_pdf(self, report_data: list[ReportRow]) -> BytesIO:
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("تقرير الحضور والانصراف", styles["Title"])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [["كود الموظف", "اسم الموظف", "القسم", "المسمى الوظيفي", "تاريخ الحضور", "وقت الحضور", "وقت الانصراف", "ساعات العمل", "الحالة", "متأخر"]]
        
        for row in report_data:
            data.append([
                row.employee_code,
                row.employee_name,
                row.department,
                row.job_title,
                row.attendance_date,
                row.check_in_time,
                row.check_out_time,
                str(row.working_hours),
                row.status,
                "نعم" if row.is_late else "لا"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
